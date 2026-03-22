"""
MySchoolio — Nepal STEM/Robotics Company Intelligence Pipeline
==============================================================
Aligned with Public_Data_Collection.txt spec (updated version).

What this script produces:
  Sheet 1 — "Company Profiles"   : Full profile per company with all spec fields
  Sheet 2 — "CRM Lead Sheet"     : Compact outreach-ready CRM view
  Sheet 3 — "Collaboration Ranks": Ranked table Level 0-5 / Bronze-Heroic
  Sheet 4 — "Daily Summary"      : Run stats and top targets

New fields vs old version:
  - Founded date
  - Team size
  - Onboarded schools count
  - Session / delivery proof
  - Collaboration level (0–5) + rank tier (Bronze → Heroic)
  - MySchoolio collaboration rationale
  - B2B model type (B2B Schools / B2C Students / Both)
  - Packages mentioned
  - Franchise/reseller model

Install:
  pip install undetected-chromedriver googlesearch-python \
              beautifulsoup4 requests pandas openpyxl lxml setuptools
"""

import re
import time
import logging
from urllib.parse import urljoin, urlparse

import requests
import pandas as pd
from bs4 import BeautifulSoup

import undetected_chromedriver as uc
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from googlesearch import search as google_search

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────────────────────
# CONFIGURATION
# ─────────────────────────────────────────────────────────────────────────────
OUTPUT_FILE       = "myschoolio_nepal_stem_leads.xlsx"
REQUEST_TIMEOUT   = 8
MAX_SUBPAGES      = 3
MAPS_SCROLL_PAUSE = 3

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s  %(levelname)-7s  %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger(__name__)


# ─────────────────────────────────────────────────────────────────────────────
# SEARCH QUERIES
# ─────────────────────────────────────────────────────────────────────────────
MAPS_QUERIES = [
    "robotics school Nepal",
    "STEM education Nepal",
    "coding school Nepal",
    "Arduino school Nepal",
    "makerspace school Nepal",
    "school robotics workshop Kathmandu",
    "electronics lab school Nepal",
    "IoT education Nepal",
    "AI school program Nepal",
    "edtech company Nepal",
    "coding bootcamp kids Nepal",
    "school science lab Nepal",
    "STEM lab setup Nepal",
    "robotics competition Nepal",
]

WEB_QUERIES = [
    '"robotics" "school" Nepal company',
    '"STEM lab" Nepal school',
    '"coding for kids" Nepal school',
    '"teacher training" STEM Nepal',
    '"school workshop" robotics Nepal',
    '"makerspace" Nepal school program',
    '"Arduino" Nepal school',
    '"Olympiad" Nepal school robotics',
    '"edtech" Nepal school',
    'Nepal "robotics club" school',
    '"AI education" Nepal school',
    'site:facebook.com "robotics" "school" Nepal',
    'site:linkedin.com/company robotics Nepal school',
]


# ─────────────────────────────────────────────────────────────────────────────
# COLLABORATION RANKING SYSTEM  (new — from spec lines 1-3)
# Level 0 Bronze  → No evidence of school collaboration
# Level 1 Silver  → Mentioned schools but no proven delivery
# Level 2 Gold    → Some school engagement, limited scale
# Level 3 Platinum→ Clear school programs, moderate scale
# Level 4 Legendary→ Strong school integration, curriculum/reporting
# Level 5 Heroic  → Deepest integration — curriculum + ops + reporting + scale
# ─────────────────────────────────────────────────────────────────────────────
COLLAB_TIERS = {
    5: "Heroic",
    4: "Legendary",
    3: "Platinum",
    2: "Gold",
    1: "Silver",
    0: "Bronze",
}

COLLAB_TIER_COLORS = {
    "Heroic":    "7030A0",   # purple
    "Legendary": "FF0000",   # red
    "Platinum":  "00B0F0",   # blue
    "Gold":      "FFD700",   # gold
    "Silver":    "A0A0A0",   # silver
    "Bronze":    "C65911",   # bronze
}

# Signals used to determine collaboration level
# Each list is checked against full scraped text
HEROIC_SIGNALS = [
    "curriculum integration", "progress report", "assessment report",
    "lesson plan", "content management", "school timetable",
    "school calendar", "teacher dashboard", "certificate",
    "school content", "batch tracking", "attendance",
]

LEGENDARY_SIGNALS = [
    "curriculum", "instructor", "assessment", "reporting",
    "school program", "school partner", "lab setup",
    "teacher training", "100+ school", "50+ school",
    "school deployment", "school implementation",
]

PLATINUM_SIGNALS = [
    "school workshop", "school visit", "school collaboration",
    "school event", "in-school", "robotics club",
    "coding for kids", "olympiad", "competition",
    "7+ school", "10+ school", "12+ school",
]

GOLD_SIGNALS = [
    "school", "student", "teacher", "education",
    "training", "workshop", "program",
]

SILVER_SIGNALS = [
    "stem", "robotics", "coding", "ai", "makerspace",
    "tech", "digital", "learning",
]

# MySchoolio collaboration rationale — what value MySchoolio adds
# Keyed by the strongest signal category detected
COLLAB_RATIONALE_MAP = {
    "curriculum": (
        "Strong fit: MySchoolio can host curriculum modules, lesson plans, and "
        "teacher handouts that align with their existing school programs."
    ),
    "report": (
        "MySchoolio can replace or enhance their manual progress/assessment "
        "reporting with structured digital dashboards per school."
    ),
    "teacher training": (
        "MySchoolio can manage training schedules, materials, and teacher "
        "certification records across all partner schools."
    ),
    "lab setup": (
        "MySchoolio can track lab activity logs, equipment usage, and "
        "school-wise STEM content access."
    ),
    "competition": (
        "MySchoolio can serve as the event management and content platform "
        "for competitions, challenges, and club activities."
    ),
    "workshop": (
        "MySchoolio can help schedule, manage, and follow up on workshops "
        "with parent updates and school-side communication."
    ),
    "default": (
        "MySchoolio can provide school content management, batch tracking, "
        "and teacher dashboards to this organization's school-facing programs."
    ),
}


# ─────────────────────────────────────────────────────────────────────────────
# RULE-BASED ANALYSIS ENGINE
# ─────────────────────────────────────────────────────────────────────────────

REJECT_WORDS = {
    "yoga", "montessori", "driving", "beauty", "salon", "fitness",
    "hospital", "clinic", "hotel", "restaurant", "travel", "trekking",
    "bank", "insurance", "pharmacy", "grocery", "real estate",
    "dance", "music", "nursery school", "coaching centre",
    "adult training", "university only", "under construction",
    "page not found",
}

REQUIRED_WORDS = {
    "robot", "stem", "cod", "school", "maker", "lab", "electron",
    "arduino", "iot", "tech", "science", "educat", "compet",
    "olympiad", "workshop", "digital", "learn",
}

CATEGORY_RULES = [
    (["robotics", "robot", "arduino", "raspberry pi"],           "Robotics Training"),
    (["stem"],                                                    "STEM Education"),
    (["coding", "programming", "code", "python", "scratch"],     "Coding for Kids"),
    (["artificial intelligence", " ai ", "machine learning"],    "AI for Students"),
    (["electronics", "circuit", "soldering", "microcontroller"], "Electronics Lab"),
    (["iot", "internet of things", "sensor"],                    "IoT Education"),
    (["workshop", "camp", "event", "training session"],          "School Workshop Provider"),
    (["teacher training", "teacher capacity", "educator"],       "Teacher Training"),
    (["competition", "olympiad", "championship", "contest"],     "Competition Organizer"),
    (["edtech", "ed-tech", "education technology", "platform"],  "EdTech Platform"),
    (["ngo", "non-profit", "nonprofit", "foundation", "trust"],  "NGO Education Partner"),
    (["lab setup", "laboratory", "science lab", "stem lab"],     "School Lab Setup"),
    (["makerspace", "maker", "fabrication", "3d print"],         "Maker Education"),
]

PROGRAM_TYPE_RULES = [
    (["robotics", "robot", "arduino"],                           "Robotics"),
    (["coding", "programming", "python", "scratch", "code"],     "Coding"),
    (["stem lab", "science lab", "laboratory"],                  "STEM Lab"),
    (["electronics", "circuit", "soldering"],                    "Electronics"),
    (["artificial intelligence", " ai ", "machine learning"],    "AI"),
    (["makerspace", "maker", "3d print", "fabrication"],         "Maker"),
    (["competition", "olympiad", "championship"],                "Competition"),
    (["workshop", "training session", "bootcamp"],               "Workshop"),
    (["teacher training", "teacher capacity", "educator"],       "Teacher Training"),
]

DELIVERY_RULES = [
    (["in-school", "in school", "school visit", "on-site",
      "school program", "school workshop"],                      "In-School"),
    (["online", "virtual", "e-learning", "digital class",
      "remote learning"],                                        "Online"),
    (["hybrid", "blended", "both online and"],                   "Hybrid"),
    (["camp", "bootcamp", "summer program", "event",
      "after-school"],                                           "Camp/Event"),
]

SEGMENT_RULES = [
    (["private school", "private schools", "boarding school"],   "Private Schools"),
    (["public school", "government school", "community school"], "Public Schools"),
    (["k-12", "k12", "kindergarten to"],                         "K-12"),
    (["secondary", "grade 9", "grade 10", "grade 11",
      "grade 12", "high school"],                               "Secondary Only"),
    (["premium", "elite", "top school"],                         "Premium Schools"),
    (["budget", "affordable", "low cost", "free program"],       "Budget Schools"),
]

B2B_RULES = [
    (["school partner", "school client", "school deployment",
      "works with school", "b2b", "institution partner"],        "B2B Schools"),
    (["student enrollment", "student signup", "join our class",
      "register as student", "b2c"],                             "B2C Students"),
]

TRUST_SIGNAL_PATTERNS = [
    (["years", "founded", "established", "since 20", "since 19"], "Years active"),
    (["testimonial", "review", "feedback", "quote"],              "Testimonials"),
    (["our team", "meet the team", "our staff", "founder"],       "Team visible"),
    (["case study", "success story", "project"],                  "Case studies"),
    (["media", "featured", "press", "news coverage"],             "Media coverage"),
    (["certified", "affiliated", "accredited"],                   "Certified/Affiliated"),
    (["client logo", "partner logo", "our school", "we work with"],"Named clients"),
]

PRICING_PATTERNS = [
    "price", "pricing", "fee", "cost", "package", "plan",
    "subscription", "rs.", "npr", "rupee", "per student",
    "per school", "per session", "quote", "contact for price",
]

PACKAGE_PATTERNS = [
    "package", "plan", "tier", "bundle", "starter", "pro",
    "enterprise", "basic plan", "premium plan",
]

FRANCHISE_PATTERNS = [
    "franchise", "reseller", "partner program", "become a partner",
    "distributor", "affiliate program", "white label",
]

# Regex patterns for structured field extraction
_SCHOOL_NAME_RE = re.compile(
    r"\b([A-Z][a-zA-Z\s]{2,40}(?:School|Academy|College|Institute|"
    r"High School|Secondary School|Boarding))\b"
)
_CONTACT_PERSON_RE = re.compile(
    r"(?:founder|ceo|director|head|manager|contact|lead by|led by)"
    r"[\s:,–\-]+([A-Z][a-z]+(?:\s[A-Z][a-z]+){1,3})",
    re.IGNORECASE,
)
_FOUNDED_RE = re.compile(
    r"(?:founded|established|since|incorporated|started)\s*(?:in\s*)?(\d{4})",
    re.IGNORECASE,
)
_TEAM_SIZE_RE = re.compile(
    r"(\d+[\+\-–]?\s*(?:to\s*\d+)?\s*(?:employees|staff|team members|people|"
    r"member|instructor|trainer))",
    re.IGNORECASE,
)
_STUDENT_COUNT_RE = re.compile(
    r"(\d[\d,]+\+?\s*(?:students|learners|kids|children)\s*(?:trained|taught|reached|enrolled)?)",
    re.IGNORECASE,
)
_SCHOOL_COUNT_RE = re.compile(
    r"(\d+\+?\s*(?:schools|institutions|colleges|centers|partner schools))",
    re.IGNORECASE,
)
_SESSION_RE = re.compile(
    r"(\d+\+?\s*(?:sessions|classes|workshops|programs|batches|projects)\s*"
    r"(?:completed|delivered|conducted|run)?)",
    re.IGNORECASE,
)
_EMAIL_RE    = re.compile(r"[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}")
_EMAIL_FP_RE = re.compile(r"\.(png|jpg|svg|woff|ttf|gif)$", re.I)
_PHONE_RE    = re.compile(r"(?:\+977[\s\-]?)?(?:9\d{9}|0\d{8,9}|01[\-\s]?\d{7})")


def _match_any(text: str, keywords: list) -> bool:
    return any(kw in text for kw in keywords)


def _collect_matches(text: str, rules: list) -> list:
    matched = []
    for keywords, label in rules:
        if _match_any(text, keywords) and label not in matched:
            matched.append(label)
    return matched


def passes_relevance_gate(name: str, extra: str = "") -> tuple:
    combined = (name + " " + extra).lower()
    for bad in REJECT_WORDS:
        if bad in combined:
            return False, f"Reject: '{bad}'"
    for good in REQUIRED_WORDS:
        if good in combined:
            return True, ""
    return False, "Reject: no STEM/school keyword found"


def _determine_collab_level(text: str) -> tuple:
    """
    Returns (level: int, tier: str, rationale: str).
    Implements the Level 0-5 / Bronze-Heroic system from the spec.
    """
    if sum(1 for s in HEROIC_SIGNALS    if s in text) >= 3:
        level = 5
    elif sum(1 for s in HEROIC_SIGNALS  if s in text) >= 1 or \
         sum(1 for s in LEGENDARY_SIGNALS if s in text) >= 3:
        level = 4
    elif sum(1 for s in LEGENDARY_SIGNALS if s in text) >= 1 or \
         sum(1 for s in PLATINUM_SIGNALS  if s in text) >= 3:
        level = 3
    elif sum(1 for s in PLATINUM_SIGNALS if s in text) >= 1 or \
         sum(1 for s in GOLD_SIGNALS     if s in text) >= 4:
        level = 2
    elif sum(1 for s in GOLD_SIGNALS    if s in text) >= 2 or \
         sum(1 for s in SILVER_SIGNALS  if s in text) >= 2:
        level = 1
    else:
        level = 0

    tier = COLLAB_TIERS[level]

    # Pick most relevant rationale
    rationale = COLLAB_RATIONALE_MAP["default"]
    for keyword, reason in COLLAB_RATIONALE_MAP.items():
        if keyword != "default" and keyword in text:
            rationale = reason
            break

    return level, tier, rationale


def analyze_content(org_name: str, full_text: str,
                    maps_cat: str = "") -> dict:
    """
    Rule-based analysis engine. Returns all fields needed for both
    the Company Profiles sheet and the CRM Lead Sheet.
    """
    text = (org_name + " " + full_text + " " + maps_cat).lower()

    # ── Categories & programs ────────────────────────────────────────────────
    categories   = _collect_matches(text, CATEGORY_RULES)
    category     = "; ".join(categories) if categories else "Education"
    prog_types   = _collect_matches(text, PROGRAM_TYPE_RULES)
    program_type = "; ".join(prog_types)
    deliveries   = _collect_matches(text, DELIVERY_RULES)
    delivery     = "; ".join(deliveries)
    segments     = _collect_matches(text, SEGMENT_RULES)
    school_seg   = "; ".join(segments)

    # ── Commercial info ───────────────────────────────────────────────────────
    pricing_visible   = "Yes" if _match_any(text, PRICING_PATTERNS) else "No"
    packages_mentioned = "Yes" if _match_any(text, PACKAGE_PATTERNS) else "No"
    franchise_model   = "Yes" if _match_any(text, FRANCHISE_PATTERNS) else "No"

    # ── B2B / B2C ────────────────────────────────────────────────────────────
    b2b_matches = _collect_matches(text, B2B_RULES)
    if len(b2b_matches) == 2:
        b2b_model = "Both"
    elif b2b_matches:
        b2b_model = b2b_matches[0]
    else:
        b2b_model = "B2B Schools"    # default assumption for school-facing orgs

    # ── Structured extractions from raw text ─────────────────────────────────
    founded_match     = _FOUNDED_RE.search(full_text)
    founded           = founded_match.group(1) if founded_match else ""

    team_match        = _TEAM_SIZE_RE.search(full_text)
    team_size         = team_match.group(0).strip() if team_match else ""

    student_match     = _STUDENT_COUNT_RE.search(full_text)
    student_count     = student_match.group(0).strip() if student_match else ""

    school_count_match = _SCHOOL_COUNT_RE.search(full_text)
    onboarded_schools = school_count_match.group(0).strip() if school_count_match else ""

    session_match     = _SESSION_RE.search(full_text)
    session_proof     = session_match.group(0).strip() if session_match else ""

    # Combine delivery proof
    delivery_proof = "; ".join(filter(None, [student_count, session_proof]))

    # Named school partners
    school_names   = _SCHOOL_NAME_RE.findall(full_text)
    seen_s         = set()
    clean_schools  = []
    for s in school_names:
        s = s.strip()
        if s and s not in seen_s and len(s) > 5:
            seen_s.add(s)
            clean_schools.append(s)
    named_partners = "; ".join(clean_schools[:8])

    # Contact person / founder
    cp_match       = _CONTACT_PERSON_RE.search(full_text)
    contact_person = cp_match.group(1).strip() if cp_match else ""

    # Trust signals
    text_trust = _collect_matches(text, TRUST_SIGNAL_PATTERNS)

    # ── School-facing signals ─────────────────────────────────────────────────
    strong_count = sum(1 for s in [
        "school partnership", "school partner", "partner school",
        "school workshop", "school event", "school program",
        "school visit", "in-school", "teacher training",
        "robotics club", "stem lab", "coding for kids",
        "olympiad", "student competition", "school competition",
        "school curriculum", "school deployment",
    ] if s in text)

    weak_count = sum(1 for w in [
        "school", "student", "teacher", "k-12", "robotics", "stem",
        "coding", "workshop", "competition", "makerspace", "edtech",
        "training", "lab", "curriculum", "olympiad", "education",
        "learning", "maker", "science", "digital",
    ] if w in text)

    if strong_count >= 1:
        school_facing = "Yes"
    elif weak_count >= 3:
        school_facing = "Yes"
    elif weak_count >= 1:
        school_facing = "Unclear"
    else:
        school_facing = "No"

    # ── Collaboration level (new) ─────────────────────────────────────────────
    collab_level, collab_tier, collab_rationale = _determine_collab_level(text)

    # ── Lead score (Section 7) ────────────────────────────────────────────────
    if school_facing == "Yes" and strong_count >= 2 and len(categories) >= 2:
        score = 5
    elif school_facing == "Yes" and (strong_count >= 1 or weak_count >= 4):
        score = 4
    elif school_facing == "Unclear" and weak_count >= 3:
        score = 3
    elif weak_count >= 2:
        score = 2
    else:
        score = 1

    # ── Red flags ─────────────────────────────────────────────────────────────
    red_flag = ""
    for trigger, reason in [
        ("yoga",               "Non-STEM: yoga"),
        ("montessori",         "Non-STEM: montessori"),
        ("adult training",     "Adult-only, not school-facing"),
        ("under construction", "Website under construction"),
        ("page not found",     "Broken website"),
    ]:
        if trigger in text:
            red_flag = reason
            break

    # ── Business note (Section 15 format) ────────────────────────────────────
    offer = program_type.split(";")[0].strip() if program_type else category.split(";")[0].strip()
    del_note     = f" via {delivery.split(';')[0].strip()}" if delivery else ""
    partner_note = f" Works with: {named_partners[:80]}." if named_partners else ""
    why_relevant = (
        f"{org_name} offers {offer}{del_note} programs in Nepal. "
        f"School-facing: {school_facing}. "
        f"{'Strong' if strong_count >= 2 else 'Some'} school engagement evidence."
        f"{partner_note}"
    )
    if red_flag:
        why_relevant = f"[Red Flag: {red_flag}] " + why_relevant

    return {
        # Core fields
        "category":              category,
        "program_type":          program_type,
        "delivery_model":        delivery,
        "school_facing":         school_facing,
        "school_segment":        school_seg,
        "b2b_model":             b2b_model,
        "pricing_visible":       pricing_visible,
        "packages_mentioned":    packages_mentioned,
        "franchise_model":       franchise_model,
        "named_school_partners": named_partners,
        "contact_person":        contact_person,
        "text_trust_signals":    "; ".join(text_trust),
        "lead_score":            score,
        "why_relevant":          why_relevant,
        "red_flag":              red_flag,
        # New fields from updated spec
        "founded":               founded,
        "team_size":             team_size,
        "onboarded_schools":     onboarded_schools,
        "student_count":         student_count,
        "session_proof":         delivery_proof,
        "collab_level":          collab_level,
        "collab_tier":           collab_tier,
        "collab_rationale":      collab_rationale,
    }


# ─────────────────────────────────────────────────────────────────────────────
# DEEP WEB SCRAPER
# ─────────────────────────────────────────────────────────────────────────────

SUBPAGE_HINTS = {
    "about", "contact", "services", "programs", "partners",
    "school", "work", "team", "mission", "what-we-do",
}

_SESSION = requests.Session()
_SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    )
})

_SOCIAL_RE = {
    "facebook":  r"https?://(?:www\.)?facebook\.com/(?!sharer)[^\s\"'<>?#]{3,}",
    "instagram": r"https?://(?:www\.)?instagram\.com/[^\s\"'<>?#]{3,}",
    "linkedin":  r"https?://(?:www\.)?linkedin\.com/company/[^\s\"'<>?#]{3,}",
    "youtube":   r"https?://(?:www\.)?youtube\.com/(?:c/|channel/|@)[^\s\"'<>?#]{3,}",
    "playstore": r"https?://play\.google\.com/store/apps/[^\s\"'<>?#]+",
}


def _fetch_page(url: str) -> tuple:
    try:
        resp = _SESSION.get(url, timeout=REQUEST_TIMEOUT)
        if resp.status_code != 200:
            return "", None
        soup = BeautifulSoup(resp.text, "lxml")
        for tag in soup(["script", "style", "noscript",
                         "header", "footer", "nav", "aside"]):
            tag.decompose()
        text = " ".join(soup.get_text(" ", strip=True).split())
        return text[:4000], soup
    except Exception as exc:
        log.debug("Fetch error %s — %s", url, exc)
        return "", None


def _subpage_urls(base_url: str, soup: BeautifulSoup) -> list:
    base  = urlparse(base_url)
    found = []
    seen  = set()
    for a in soup.find_all("a", href=True):
        href = a["href"].lower().strip()
        if any(hint in href for hint in SUBPAGE_HINTS):
            full = urljoin(base_url, a["href"])
            p    = urlparse(full)
            if p.netloc == base.netloc and full not in seen:
                found.append(full)
                seen.add(full)
        if len(found) >= MAX_SUBPAGES:
            break
    return found


def deep_scrape(url: str) -> dict:
    result = {k: "" for k in ["text", "email", "phone", "facebook",
                               "instagram", "linkedin", "youtube", "playstore"]}
    if not url:
        return result

    home_text, home_soup = _fetch_page(url)
    if not home_soup:
        return result

    all_text = home_text
    for sp_url in _subpage_urls(url, home_soup)[:MAX_SUBPAGES]:
        sp_text, _ = _fetch_page(sp_url)
        if sp_text:
            all_text += " " + sp_text

    result["text"] = all_text[:10000]

    # Email
    emails = _EMAIL_RE.findall(all_text)
    result["email"] = next(
        (e for e in emails if not _EMAIL_FP_RE.search(e)), ""
    )

    # Phone (new — from spec)
    phone_match = _PHONE_RE.search(all_text)
    result["phone"] = phone_match.group(0).strip() if phone_match else ""

    # Social links
    raw_html = str(home_soup)
    for platform, pattern in _SOCIAL_RE.items():
        m = re.search(pattern, raw_html)
        if m:
            result[platform] = m.group(0).rstrip("/,")

    return result


# ─────────────────────────────────────────────────────────────────────────────
# SOCIAL LINK VALIDATOR
# ─────────────────────────────────────────────────────────────────────────────

def _url_alive(url: str) -> bool:
    if not url:
        return False
    try:
        r = _SESSION.head(url, timeout=6, allow_redirects=True)
        if r.status_code < 400:
            return True
        r = _SESSION.get(url, timeout=6, stream=True)
        return r.status_code < 400
    except Exception:
        return False


def validated_social(raw: dict) -> dict:
    out = dict(raw)
    for p in ["facebook", "instagram", "linkedin", "youtube", "playstore"]:
        if out.get(p) and not _url_alive(out[p]):
            log.debug("  Dead link cleared: %s", p)
            out[p] = ""
    return out


# ─────────────────────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _priority(score: int) -> str:
    return "High" if score >= 4 else ("Medium" if score == 3 else "Low")


def _trust_signals(rating: str, reviews: str,
                   social: dict, text_trust: str) -> str:
    parts = []
    if rating:   parts.append(f"Rating: {rating}")
    if reviews:  parts.append(f"Reviews: {reviews}")
    for p in ["facebook", "instagram", "linkedin", "youtube"]:
        if social.get(p):
            parts.append(f"{p.title()} active")
    if text_trust:
        parts.extend(text_trust.split("; "))
    return "; ".join(dict.fromkeys(parts))


def _build_lead(lead_id: int, name: str, city: str,
                maps_phone: str, rating: str, reviews: str,
                website: str, maps_cat: str,
                deep: dict, analysis: dict, source3: str) -> dict:

    social   = validated_social(deep)
    score    = analysis["lead_score"]
    priority = _priority(score)
    phone    = deep.get("phone") or maps_phone

    return {
        # Identity
        "Lead ID":               f"L{lead_id:04d}",
        "Organization Name":     name,
        "Website":               website,
        "Facebook":              social.get("facebook", ""),
        "LinkedIn":              social.get("linkedin", ""),
        "Instagram":             social.get("instagram", ""),
        "YouTube":               social.get("youtube", ""),
        "City":                  city,
        "Country":               "Nepal",
        # Company profile fields (new)
        "Founded":               analysis["founded"],
        "Team Size":             analysis["team_size"],
        "Onboarded Schools":     analysis["onboarded_schools"],
        "Students Reached":      analysis["student_count"],
        "Session / Delivery Proof": analysis["session_proof"],
        # Classification
        "Category":              analysis["category"],
        "Teaching Field":        analysis["program_type"],
        "Focused Area":          analysis["category"],
        "Main Offer":            (deep.get("text") or "")[:200],
        "School Facing?":        analysis["school_facing"],
        "B2B Model":             analysis["b2b_model"],
        "School Segment":        analysis["school_segment"],
        "Program Type":          analysis["program_type"],
        "Delivery Model":        analysis["delivery_model"],
        # Commercial
        "Pricing Visible?":      analysis["pricing_visible"],
        "Packages Mentioned?":   analysis["packages_mentioned"],
        "Franchise/Reseller?":   analysis["franchise_model"],
        "Named School Partners": analysis["named_school_partners"],
        # Contact
        "Public Contact Email":  deep.get("email", ""),
        "Public Contact Phone":  phone,
        "Contact Person":        analysis["contact_person"],
        "Last Active Date":      "",           # manual
        # Trust
        "Trust Signals":         _trust_signals(
                                     rating, reviews, social,
                                     analysis.get("text_trust_signals", "")
                                 ),
        "Rating":                rating,
        "Review Count":          reviews,
        # Scoring
        "Lead Score":            score,
        "Priority":              priority,
        # Collaboration rank (new)
        "Collaboration Level":   analysis["collab_level"],
        "Collaboration Tier":    analysis["collab_tier"],
        "MySchoolio Fit":        analysis["collab_rationale"],
        # Notes
        "Why Relevant":          analysis["why_relevant"],
        "Notes":                 analysis["why_relevant"],
        # Sources
        "Source 1":              website,
        "Source 2":              social.get("facebook", ""),
        "Source 3":              source3,
        "Maps Category":         maps_cat,
    }


# ─────────────────────────────────────────────────────────────────────────────
# GOOGLE MAPS SCRAPER
# ─────────────────────────────────────────────────────────────────────────────

def _init_driver() -> uc.Chrome:
    opts = uc.ChromeOptions()
    # opts.add_argument("--headless=new")
    opts.add_argument("--no-sandbox")
    opts.add_argument("--disable-dev-shm-usage")
    opts.add_argument("--window-size=1280,900")
    return uc.Chrome(options=opts)


def scrape_maps(driver, wait, query, seen, data, counter):
    log.info("[Maps] %s", query)
    maps_url = "https://www.google.com/maps/search/" + query.replace(" ", "+")
    driver.get(maps_url)
    time.sleep(5)

    try:
        feed = driver.find_element(By.CSS_SELECTOR, 'div[role="feed"]')
    except Exception:
        log.warning("  No feed — skipping: %s", query)
        return

    last_h = 0
    while True:
        driver.execute_script(
            "arguments[0].scrollTop = arguments[0].scrollHeight", feed)
        time.sleep(MAPS_SCROLL_PAUSE)
        new_h = driver.execute_script("return arguments[0].scrollHeight", feed)
        if new_h == last_h:
            break
        last_h = new_h

    listings = driver.find_elements(By.CSS_SELECTOR, "div.Nv2PK")
    log.info("  Raw listings: %d", len(listings))

    for i in range(len(listings)):
        listings = driver.find_elements(By.CSS_SELECTOR, "div.Nv2PK")
        if i >= len(listings):
            break
        try:
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", listings[i])
            time.sleep(0.8)
            try:
                listings[i].click()
            except Exception:
                driver.execute_script("arguments[0].click();", listings[i])
            time.sleep(3)
            name = wait.until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "h1.DUwDvf"))
            ).text.strip()
        except Exception:
            continue

        if not name or name in seen:
            continue

        try:
            maps_cat = driver.find_element(By.CSS_SELECTOR, "button.DkEaL").text
        except Exception:
            maps_cat = ""

        keep, reason = passes_relevance_gate(name, maps_cat)
        if not keep:
            log.debug("  SKIP %s — %s", name, reason)
            continue

        seen.add(name)
        time.sleep(0.5)

        try:
            address = driver.find_element(
                By.CSS_SELECTOR, 'button[data-item-id="address"]').text
        except Exception:
            address = ""

        city_guess = address.split(",")[-1].strip() if address else "Nepal"

        try:
            phone = driver.find_element(
                By.CSS_SELECTOR, 'button[data-item-id^="phone"]').text
        except Exception:
            phone = ""

        try:
            website = driver.find_element(
                By.CSS_SELECTOR, 'a[data-item-id="authority"]'
            ).get_attribute("href") or ""
        except Exception:
            website = ""

        try:
            rating = driver.find_element(By.CSS_SELECTOR, "span.F7nice").text
        except Exception:
            rating = ""

        try:
            reviews = driver.find_element(
                By.CSS_SELECTOR, "span.F7nice + span").text
        except Exception:
            reviews = ""

        if not website:
            website = _find_website_fallback(name)

        log.info("  Deep scraping: %s", website or "(none)")
        deep = deep_scrape(website)

        keep2, reason2 = passes_relevance_gate(name, deep.get("text", "")[:300])
        if not keep2:
            log.debug("  SKIP (text) %s — %s", name, reason2)
            seen.discard(name)
            continue

        analysis = analyze_content(name, deep.get("text", ""), maps_cat)
        lead     = _build_lead(counter[0], name, city_guess, phone, rating,
                               reviews, website, maps_cat, deep, analysis, maps_url)
        data.append(lead)
        counter[0] += 1
        log.info("  ✓ %s | Score:%s | Level:%s %s",
                 name, lead["Lead Score"],
                 lead["Collaboration Level"], lead["Collaboration Tier"])


# ─────────────────────────────────────────────────────────────────────────────
# WEB SEARCH FALLBACK
# ─────────────────────────────────────────────────────────────────────────────

_SOCIAL_DOMAINS = {
    "facebook.com", "linkedin.com", "youtube.com",
    "twitter.com", "instagram.com", "wikipedia.org",
}
_SKIP_DOMAINS = {
    "wikipedia.org", "youtube.com", "tripadvisor",
    "justdial", "facebook.com/groups", "linkedin.com/in/",
}


def _google_search_safe(query: str, num: int = 10) -> list:
    """
    Wrapper around googlesearch.search() that handles both old and new
    versions of the library gracefully.

    Old API (googlesearch-python < 1.2):  search(q, num=N, stop=N, pause=2)
    New API (googlesearch-python >= 1.2):  search(q, num_results=N)
    """
    try:
        # Try new API first
        return list(google_search(query, num_results=num))
    except TypeError:
        pass
    try:
        # Fall back to old API
        return list(google_search(query, num=num, stop=num, pause=2))
    except Exception as exc:
        log.debug("google_search failed: %s", exc)
        return []


def _find_website_fallback(org_name: str) -> str:
    for url in _google_search_safe(f"{org_name} Nepal official site", num=5):
        if not any(d in url for d in _SOCIAL_DOMAINS):
            return url
    return ""


def scrape_web_search(query, seen, data, counter):
    log.info("[Web] %s", query)
    results = _google_search_safe(query, num=10)
    if not results:
        log.warning("  No results returned for: %s", query)
        return

    for url in results:
        if any(d in url for d in _SKIP_DOMAINS):
            continue
        parsed    = urlparse(url)
        prov_name = parsed.netloc.replace("www.", "").split(".")[0].title()
        if prov_name in seen:
            continue

        deep = deep_scrape(url)
        keep, reason = passes_relevance_gate(prov_name, deep.get("text", "")[:300])
        if not keep:
            continue

        try:
            _, soup = _fetch_page(url)
            if soup and soup.title:
                t = soup.title.get_text(strip=True)[:80]
                if t:
                    prov_name = t
        except Exception:
            pass

        if prov_name in seen:
            continue
        seen.add(prov_name)

        analysis = analyze_content(prov_name, deep.get("text", ""))
        lead     = _build_lead(
            counter[0], prov_name, "Nepal", "", "", "",
            url, "Web Search", deep, analysis,
            f"https://www.google.com/search?q={query.replace(' ', '+')}"
        )
        data.append(lead)
        counter[0] += 1
        log.info("  ✓ %s | Score:%s | Level:%s %s",
                 prov_name, lead["Lead Score"],
                 lead["Collaboration Level"], lead["Collaboration Tier"])
        time.sleep(1.5)


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL OUTPUT — 4 sheets
# ─────────────────────────────────────────────────────────────────────────────

def _apply_header(ws, font_color="FFFFFF", fill_color="1F4E79"):
    fill = PatternFill("solid", start_color=fill_color, end_color=fill_color)
    font = Font(bold=True, color=font_color, name="Arial", size=10)
    for cell in ws[1]:
        cell.font      = font
        cell.fill      = fill
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
    ws.row_dimensions[1].height = 30


def _apply_body(ws, priority_col_name="Priority"):
    BODY_FONT = Font(name="Arial", size=9)
    HIGH_FILL = PatternFill("solid", start_color="C6EFCE", end_color="C6EFCE")
    MED_FILL  = PatternFill("solid", start_color="FFEB9C", end_color="FFEB9C")
    LOW_FILL  = PatternFill("solid", start_color="FFC7CE", end_color="FFC7CE")

    priority_col = next(
        (i for i, c in enumerate(ws[1], 1) if c.value == priority_col_name),
        None
    )

    for row in ws.iter_rows(min_row=2):
        pval = row[priority_col - 1].value if priority_col else ""
        fill = (HIGH_FILL if pval == "High"
                else MED_FILL if pval == "Medium"
                else LOW_FILL)
        for cell in row:
            cell.font      = BODY_FONT
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.fill      = fill


def _set_col_widths(ws, widths: dict):
    for col in ws.columns:
        letter = get_column_letter(col[0].column)
        ws.column_dimensions[letter].width = widths.get(col[0].value, 18)


def save_excel(df: pd.DataFrame, path: str) -> None:
    """
    Writes all 4 sheets in a single ExcelWriter pass to avoid the
    sheet-overwrite bug that occurred when calling df.to_excel() twice
    on the same path.
    """

    # ── Prepare sub-dataframes ────────────────────────────────────────────────
    profile_cols = [
        "Lead ID", "Organization Name", "Website", "Facebook", "LinkedIn",
        "Instagram", "YouTube", "City", "Country",
        "Founded", "Team Size", "Onboarded Schools", "Students Reached",
        "Session / Delivery Proof",
        "Category", "Teaching Field", "Focused Area",
        "School Facing?", "B2B Model", "School Segment",
        "Program Type", "Delivery Model",
        "Pricing Visible?", "Packages Mentioned?", "Franchise/Reseller?",
        "Named School Partners",
        "Public Contact Email", "Public Contact Phone", "Contact Person",
        "Last Active Date", "Trust Signals", "Rating", "Review Count",
        "Lead Score", "Priority",
        "Collaboration Level", "Collaboration Tier", "MySchoolio Fit",
        "Why Relevant", "Source 1", "Source 2", "Source 3", "Maps Category",
    ]
    crm_cols = [
        "Lead ID", "Organization Name", "City",
        "Public Contact Email", "Public Contact Phone", "Contact Person",
        "Founded", "Team Size", "Onboarded Schools", "Students Reached",
        "Teaching Field", "Delivery Model", "School Facing?", "B2B Model",
        "Lead Score", "Priority",
        "Collaboration Level", "Collaboration Tier", "MySchoolio Fit",
        "Facebook", "LinkedIn", "Website",
    ]
    rank_data_cols = [
        "Organization Name", "Collaboration Level", "Collaboration Tier",
        "Lead Score", "Onboarded Schools", "Students Reached",
        "MySchoolio Fit", "Contact Person",
        "Public Contact Email", "Public Contact Phone", "Website",
    ]

    df_profile = df[[c for c in profile_cols if c in df.columns]].copy()
    df_crm     = df[[c for c in crm_cols     if c in df.columns]].copy()
    df_rank    = df[[c for c in rank_data_cols if c in df.columns]].copy()
    df_rank    = df_rank.sort_values(
        ["Collaboration Level", "Lead Score"], ascending=[False, False]
    ).reset_index(drop=True)

    # ── Write all sheets in one pass ──────────────────────────────────────────
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df_profile.to_excel(writer, index=False, sheet_name="Company Profiles")
        df_crm.to_excel(writer,     index=False, sheet_name="CRM Lead Sheet")
        # Rank and Summary sheets are built manually below via openpyxl
        # Write a placeholder so openpyxl knows the sheets exist
        pd.DataFrame().to_excel(writer, sheet_name="Collaboration Ranks")
        pd.DataFrame().to_excel(writer, sheet_name="Daily Summary")

    # ── Re-open for styling ───────────────────────────────────────────────────
    wb = load_workbook(path)

    # ── Sheet 1: Company Profiles ─────────────────────────────────────────────
    ws1 = wb["Company Profiles"]
    _apply_header(ws1)
    _apply_body(ws1)
    _set_col_widths(ws1, {
        "Lead ID": 8, "Organization Name": 28, "Website": 28,
        "Facebook": 25, "LinkedIn": 22, "Instagram": 22, "YouTube": 22,
        "City": 12, "Country": 8, "Founded": 8, "Team Size": 14,
        "Onboarded Schools": 16, "Students Reached": 16,
        "Session / Delivery Proof": 22,
        "Category": 25, "Teaching Field": 25, "Focused Area": 25,
        "School Facing?": 12, "B2B Model": 14, "School Segment": 20,
        "Program Type": 22, "Delivery Model": 16,
        "Pricing Visible?": 12, "Packages Mentioned?": 14,
        "Franchise/Reseller?": 14, "Named School Partners": 30,
        "Public Contact Email": 26, "Public Contact Phone": 16,
        "Contact Person": 18, "Last Active Date": 14,
        "Trust Signals": 25, "Rating": 8, "Review Count": 12,
        "Lead Score": 10, "Priority": 10,
        "Collaboration Level": 14, "Collaboration Tier": 14,
        "MySchoolio Fit": 40, "Why Relevant": 38,
        "Source 1": 30, "Source 2": 28, "Source 3": 28,
        "Maps Category": 20,
    })
    ws1.freeze_panes    = "A2"
    ws1.auto_filter.ref = ws1.dimensions

    # ── Sheet 2: CRM Lead Sheet ───────────────────────────────────────────────
    ws_crm = wb["CRM Lead Sheet"]
    _apply_header(ws_crm, fill_color="1A5276")
    _apply_body(ws_crm)
    _set_col_widths(ws_crm, {
        "Lead ID": 8, "Organization Name": 28, "City": 12,
        "Public Contact Email": 26, "Public Contact Phone": 16,
        "Contact Person": 18, "Founded": 8, "Team Size": 14,
        "Onboarded Schools": 16, "Students Reached": 16,
        "Teaching Field": 25, "Delivery Model": 16,
        "School Facing?": 12, "B2B Model": 14,
        "Lead Score": 10, "Priority": 10,
        "Collaboration Level": 14, "Collaboration Tier": 14,
        "MySchoolio Fit": 42, "Facebook": 25,
        "LinkedIn": 22, "Website": 28,
    })
    ws_crm.freeze_panes    = "A2"
    ws_crm.auto_filter.ref = ws_crm.dimensions

    # ── Sheet 3: Collaboration Ranks (rebuilt manually) ───────────────────────
    ws_rank = wb["Collaboration Ranks"]
    # Clear the placeholder content written by pd.DataFrame().to_excel()
    for row in ws_rank.iter_rows():
        for cell in row:
            cell.value = None

    rank_headers = ["Rank"] + rank_data_cols
    ws_rank.append(rank_headers)
    _apply_header(ws_rank, fill_color="4A235A")

    for idx, (_, row) in enumerate(df_rank.iterrows(), 1):
        tier       = str(row.get("Collaboration Tier", "") if hasattr(row, "get")
                         else row["Collaboration Tier"] if "Collaboration Tier" in row.index else "")
        color      = COLLAB_TIER_COLORS.get(tier, "FFFFFF")
        fill       = PatternFill("solid", start_color=color, end_color=color)
        dark_bg    = tier in ("Heroic", "Legendary", "Bronze")
        font_color = "FFFFFF" if dark_bg else "000000"

        row_values = [idx] + [
            row[c] if c in row.index else "" for c in rank_data_cols
        ]
        ws_rank.append(row_values)
        for cell in ws_rank[ws_rank.max_row]:
            cell.fill      = fill
            cell.font      = Font(name="Arial", size=9, color=font_color)
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    _set_col_widths(ws_rank, {
        "Rank": 6, "Organization Name": 28, "Collaboration Level": 14,
        "Collaboration Tier": 14, "Lead Score": 10,
        "Onboarded Schools": 16, "Students Reached": 16,
        "MySchoolio Fit": 42, "Contact Person": 18,
        "Public Contact Email": 26, "Public Contact Phone": 16,
        "Website": 28,
    })
    ws_rank.freeze_panes    = "A2"
    ws_rank.auto_filter.ref = ws_rank.dimensions

    # ── Sheet 4: Daily Summary (rebuilt manually) ─────────────────────────────
    ws_sum = wb["Daily Summary"]
    for row in ws_sum.iter_rows():
        for cell in row:
            cell.value = None

    tot    = len(df)
    ver    = int((df["Lead Score"] >= 4).sum())
    top5   = df.nlargest(5, "Lead Score")[
        ["Organization Name", "Lead Score", "Collaboration Tier", "MySchoolio Fit"]
    ]
    heroic_count    = int((df["Collaboration Tier"] == "Heroic").sum())
    legendary_count = int((df["Collaboration Tier"] == "Legendary").sum())
    platinum_count  = int((df["Collaboration Tier"] == "Platinum").sum())

    summary_rows = [
        ["MySchoolio — Nepal STEM/Robotics Lead Intelligence Report"],
        [],
        ["Metric",                          "Value"],
        ["Total leads saved",               tot],
        ["Verified leads (score ≥ 4)",      ver],
        ["High-priority leads",             int((df["Priority"] == "High").sum())],
        ["Medium-priority leads",           int((df["Priority"] == "Medium").sum())],
        [],
        ["Collaboration Tiers"],
        ["Heroic (Level 5)",                heroic_count],
        ["Legendary (Level 4)",             legendary_count],
        ["Platinum (Level 3)",              platinum_count],
        [],
        ["Top 5 Strongest Leads for MySchoolio Outreach"],
        ["Organization Name", "Lead Score", "Collab Tier", "MySchoolio Fit"],
    ]
    for _, r in top5.iterrows():
        summary_rows.append([
            r["Organization Name"], r["Lead Score"],
            r["Collaboration Tier"], r["MySchoolio Fit"],
        ])
    summary_rows += [
        [],
        ["Next Steps"],
        ["1. Prioritize Heroic and Legendary tier leads for first outreach."],
        ["2. Fill 'Last Active Date' and 'Contact Person' manually."],
        ["3. Score-3/Platinum leads need a validation call before pitching."],
        ["4. Use the CRM Lead Sheet tab for outreach tracking."],
    ]
    for row in summary_rows:
        ws_sum.append(row)

    ws_sum["A1"].font  = Font(bold=True, size=13, name="Arial")
    ws_sum["A3"].font  = Font(bold=True, name="Arial")
    ws_sum["A9"].font  = Font(bold=True, name="Arial")
    ws_sum["A14"].font = Font(bold=True, name="Arial")
    ws_sum.column_dimensions["A"].width = 42
    ws_sum.column_dimensions["B"].width = 14
    ws_sum.column_dimensions["C"].width = 16
    ws_sum.column_dimensions["D"].width = 55

    # ── Reorder sheets safely ─────────────────────────────────────────────────
    desired_order = [
        "Company Profiles", "CRM Lead Sheet",
        "Collaboration Ranks", "Daily Summary",
    ]
    for target_idx, sheet_name in enumerate(desired_order):
        if sheet_name in wb.sheetnames:
            current_idx = wb.sheetnames.index(sheet_name)
            wb.move_sheet(sheet_name, offset=target_idx - current_idx)

    wb.save(path)
    log.info("Excel saved → %s", path)


# ─────────────────────────────────────────────────────────────────────────────
# MAIN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def main():
    data    = []
    seen    = set()
    counter = [1]

    # Pass 1 — Google Maps
    driver = _init_driver()
    wait   = WebDriverWait(driver, 12)
    try:
        for q in MAPS_QUERIES:
            scrape_maps(driver, wait, q, seen, data, counter)
            time.sleep(2)
    finally:
        driver.quit()
    log.info("Maps pass done. Leads: %d", len(data))

    # Pass 2 — Web search fallback
    for q in WEB_QUERIES:
        scrape_web_search(q, seen, data, counter)
        time.sleep(3)
    log.info("Web pass done. Total: %d", len(data))

    df = pd.DataFrame(data)
    if df.empty:
        log.warning("No leads collected. Check queries and internet connection.")
        return

    df = df.drop_duplicates(subset=["Organization Name"])
    df = df.sort_values(
        ["Collaboration Level", "Lead Score"], ascending=[False, False]
    ).reset_index(drop=True)

    save_excel(df, OUTPUT_FILE)

    log.info("─" * 58)
    log.info("DONE  →  %s", OUTPUT_FILE)
    log.info("  Total saved     : %d", len(df))
    log.info("  Score ≥ 4       : %d", int((df["Lead Score"] >= 4).sum()))
    log.info("  Heroic/Legendary: %d",
             int((df["Collaboration Tier"].isin(["Heroic", "Legendary"])).sum()))
    log.info("─" * 58)


if __name__ == "__main__":
    main()