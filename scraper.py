import time
import re
import requests
import pandas as pd

from bs4 import BeautifulSoup

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager

from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter


# SEARCH SETTINGS
keywords = [
    "education consultancy",
    # "IT training institute",
    # "digital marketing agency"
]

cities = [
    "kathmandu",
    "lalitpur",
    "bhaktapur",
    "pokhara"
]


driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()))
wait = WebDriverWait(driver, 10)

data = []


def extract_contact_info(url):

    email = ""
    facebook = ""
    instagram = ""

    try:

        response = requests.get(url, timeout=5)
        soup = BeautifulSoup(response.text, "html.parser")

        emails = re.findall(
            r"[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Za-z]{2,}",
            response.text
        )

        if emails:
            email = emails[0]

        for link in soup.find_all("a", href=True):

            href = link["href"]

            if "facebook.com" in href and facebook == "":
                facebook = href

            if "instagram.com" in href and instagram == "":
                instagram = href

    except:
        pass

    return email, facebook, instagram


for keyword in keywords:
    for city in cities:

        query = f"{keyword} {city}"

        print("\nSearching:", query)

        driver.get(f"https://www.google.com/maps/search/{query}")

        time.sleep(5)

        scrollable_div = driver.find_element(By.CSS_SELECTOR, 'div[role="feed"]')

        last_height = 0

        while True:

            driver.execute_script(
                "arguments[0].scrollTop = arguments[0].scrollHeight",
                scrollable_div
            )

            time.sleep(3)

            new_height = driver.execute_script(
                "return arguments[0].scrollHeight",
                scrollable_div
            )

            if new_height == last_height:
                break

            last_height = new_height


        listings = driver.find_elements(By.CSS_SELECTOR, "div.Nv2PK")

        print("Businesses found:", len(listings))


        for i in range(len(listings)):

            listings = driver.find_elements(By.CSS_SELECTOR, "div.Nv2PK")
            
            # Check to ensure index hasn't gone out of bounds if DOM changes
            if i >= len(listings):
                break
                
            listing = listings[i]

            try:
                # Scroll into view to ensure it can be clicked without being obstructed
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", listing)
                time.sleep(1)

                # Try standard click, fallback to script click if obstructed
                try:
                    listing.click()
                except:
                    driver.execute_script("arguments[0].click();", listing)

                # Wait for the left panel to update with the new business data.
                # This delay is crucial to prevent reading the previously clicked business.
                time.sleep(3)

                name = wait.until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "h1.DUwDvf"))
                ).text

            except Exception as e:
                name = ""

            # Small additional buffer for other details to load
            time.sleep(1)

            try:
                address = driver.find_element(
                    By.CSS_SELECTOR,
                    'button[data-item-id="address"]'
                ).text
            except:
                address = ""

            try:
                phone = driver.find_element(
                    By.CSS_SELECTOR,
                    'button[data-item-id^="phone"]'
                ).text
            except:
                phone = ""

            try:
                website = driver.find_element(
                    By.CSS_SELECTOR,
                    'a[data-item-id="authority"]'
                ).get_attribute("href")
            except:
                website = ""

            try:
                rating = driver.find_element(
                    By.CSS_SELECTOR,
                    "span.F7nice"
                ).text
            except:
                rating = ""

            try:
                reviews = driver.find_element(
                    By.CSS_SELECTOR,
                    "span.F7nice + span"
                ).text
            except:
                reviews = ""

            email = ""
            facebook = ""
            instagram = ""

            if website:

                email, facebook, instagram = extract_contact_info(website)

            data.append([
                keyword,
                city,
                name,
                address,
                phone,
                website,
                rating,
                reviews,
                email,
                facebook,
                instagram
            ])

            print("Collected:", name)


driver.quit()


# CREATE DATAFRAME
df = pd.DataFrame(
    data,
    columns=[
        "Category",
        "City",
        "Business Name",
        "Address",
        "Phone",
        "Website",
        "Rating",
        "Review Count",
        "Email",
        "Facebook",
        "Instagram"
    ]
)


# REMOVE DUPLICATES
df = df.drop_duplicates(subset=["Business Name"])


# SAVE TO EXCEL
file_name = "education_consultancy_leads.xlsx"
df.to_excel(file_name, index=False)


# FORMAT EXCEL
wb = load_workbook(file_name)
ws = wb.active


# Bold header
for cell in ws[1]:
    cell.font = Font(bold=True)


# Wrap text
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)


# Auto column width
for column in ws.columns:
    max_length = 0
    column_letter = get_column_letter(column[0].column)

    for cell in column:
        try:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        except:
            pass

    ws.column_dimensions[column_letter].width = min(max_length + 5, 50)


# Freeze header
ws.freeze_panes = "A2"


wb.save(file_name)

print("\nScraping complete. File saved:", file_name)